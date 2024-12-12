import os
from datetime import datetime

from copy_file import copy_file
from display_message import display_message
from openpyxl import load_workbook

from log_writer import log_writer
from mappings import paths


def save_planilha_resultado_final(nome, ultimo_id, time_casa, placar_casa, time_visitante, placar_visitante, codigo_partida):
    # print("save_planilha_resultado_final")
    local_path = paths("file_server_local")
    gdrive_path = paths("file_server_gdrive")

    try:
        planilha_gdrive_path = gdrive_path + nome + ".xlsx"
        planilha_local_path = local_path + nome

        # Carregar a planilha existente
        if os.path.exists(planilha_gdrive_path):
            workbook = load_workbook(planilha_gdrive_path)
        else:
            message = f"Erro: Planilha {planilha_gdrive_path} não encontrada."
            display_message(message)
            log_writer(message)
            return

        # Selecionar as abas
        sheet_name_resultados = "Resultados"  # Defina o nome da aba que você deseja selecionar

        if sheet_name_resultados in workbook.sheetnames:
            sheet_resultados = workbook[sheet_name_resultados]
        else:
            message = f"Aba {sheet_name_resultados} não encontrada. Criando nova aba."
            display_message(message)
            log_writer(message)

            sheet_resultados = workbook.create_sheet(title=sheet_name_resultados)

        # Encontrar a próxima linha vazia
        row = int(ultimo_id) + 1

        # Gravar os dados na próxima linha
        current_time = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        sheet_resultados[f"B{row}"] = time_casa
        sheet_resultados[f"C{row}"] = placar_casa
        sheet_resultados[f"D{row}"] = placar_visitante
        sheet_resultados[f"E{row}"] = time_visitante
        sheet_resultados[f"G{row}"] = current_time
        sheet_resultados[f"H{row}"] = codigo_partida

        # Salvar a planilha
        workbook.save(planilha_gdrive_path)
        # display_message(f"Resultados finais gravados na planilha {planilha_gdrive_path} para {time_casa} x {time_visitante}")

        copy_file(planilha_gdrive_path, planilha_local_path)

    except Exception as e:
        message = f"Erro ao salvar o resultado na planilha: {str(e)}"
        display_message(message)
        log_writer(message)
